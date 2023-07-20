from numpy import *
from tablas import TABLA_17_19_1,  TABLA_17_20_A, TABLA_17_20_B, TABLA_17_20_C, TABLA_17_20_C2, TABLA_17_22, TABLA_17_23, Tabla_17_15
from openpyxl.styles import Alignment, Border, Side
from openpyxl import load_workbook,Workbook
Circuito = True

# Funcion para encontrar la longitud de paso
def long_paso(Cp, N1, N2):
    long_paso = 2 * Cp + (N1 + N2)/2 + ((N2 - N1)**2)/(4 * (pi)**2 * Cp)
    return long_paso

# Función para calcular la longitud de la cadena
def longitud(Lp, p):
    long = Lp * p
    return long

#Funcion para distancia de centros
def distan_centro(Cp, p):
    dis_centro = Cp * p
    return dis_centro

#Funciones para calcular el Htab requerido
def calculo_Htab_requerido(H_mon, k1, k2, n_d, K_s):
    Htab = (n_d * K_s * H_mon)/(k1 * k2)
    return Htab

#Funcion iterar el numero de hileras
def hileras(valor):
    n_hileras = TABLA_17_23[valor]["K2"]
    return n_hileras

#Funcion para encontrar k1
def factor_k1(valor):
    if(valor > 20):
        k1 = (valor/17)**1.08
        return k1
    else:
        for i in range(len(TABLA_17_22)):
            if TABLA_17_22[i]["Número de dientes de catarina impulsora"] == valor:
                return TABLA_17_22[i]["Potencia preextremo K1"]

# Función para obtener el factor de servicio Ks según el tipo de máquina y par de torsión
def factor_servicio(tipo):
    while True:
        try:
            maquina = int(input("\nElija maquinaria impulsada:\n1. Uniforme.\n2. Impacto ligero.\n3. Impacto medio.\n4. Impacto pesado\nElección: "))
            if maquina != 1 and maquina != 2 and maquina != 3 and maquina != 4:
                print("Ingrese una opción válida")
            else: 
                break
        except (ValueError):
            print("Ingrese una opción válida")
        
    if tipo == 1:
        if maquina == 1:
            Ks = Tabla_17_15[0][0]["Ks"]
        elif maquina == 2:
            Ks = Tabla_17_15[0][1]["Ks"]
        elif maquina == 3:
            Ks = Tabla_17_15[0][2]["Ks"]
        elif maquina == 4:
            Ks = Tabla_17_15[0][3]["Ks"]
    elif tipo == 2:
        if maquina == 1:
            Ks = Tabla_17_15[1][0]["Ks"]
        elif maquina == 2:
            Ks = Tabla_17_15[1][1]["Ks"]
        elif maquina == 3:
            Ks = Tabla_17_15[1][2]["Ks"]
        elif maquina == 4:
            Ks = Tabla_17_15[1][3]["Ks"]
    return Ks

#Funcion para encontrar el Htab 
def buscar_Htab_tablas(tabla, rpm, H_nom, lubri,index):
    xd = None
    vectorcito = []
    for i in range(len(tabla)):
        if rpm <= tabla[i]["RPM"]:
            rpm_actual = tabla[i]["RPM"]
            break
    for j in range(len(tabla)):
        if rpm_actual == tabla[j]["RPM"]:
            if H_nom <= tabla[j]["HP"]:
                xd = any
                vectorcito.append({"# Hileras": index, "H_REQ": "{:.3f}".format(H_nom),"H_TAB": tabla[j]["HP"],"Lubri": lubri, "ANSI":tabla[j]["TIPO"]})

    if xd == None:
        return []
    else:     
        return vectorcito

#Funcion para el buscar el Htab en todas las tablas
def H_TAB(TABLA_17_20_A, TABLA_17_20_B, TABLA_17_20_C, TABLA_17_20_C2, rpm, H_nom, index):
    a = buscar_Htab_tablas(TABLA_17_20_A, rpm, H_nom, "TIPO A", index)
    b = buscar_Htab_tablas(TABLA_17_20_B, rpm, H_nom, "TIPO B", index)
    c = buscar_Htab_tablas(TABLA_17_20_C, rpm, H_nom, "TIPO C", index)
    c2= buscar_Htab_tablas(TABLA_17_20_C2, rpm, H_nom, "TIPO C2", index)

    if a is not None or b is not None or c is not None or c2 is not None:
        return a + b + c + c2

#Funcion para imprimir las tablas encontradas
def imprimir_tabla(tabla):
    headers = list(tabla[0].keys())
    header_format = " | ".join(["{:<10}".format(h) for h in headers])
    separator = "-" * len(header_format)
    print(header_format)
    print(separator)
    for item in tabla:
        row_format = " | ".join(["{:<10}".format(item[key]) for key in headers])
        print(row_format)

#Funcion para guardar los datos en un Excel
def guardar_datos_en_tabla(Tip, paso, Ancho, Resis, Peso_prom, Diametro_resis, E_H_M, Lp, lon, cen):
    try:
        wb = load_workbook("Tabla.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    # Encabezados de la tabla (si es la primera vez que se guarda)
    if ws.cell(row=1, column=1).value is None:
        ws.cell(row=1, column=1, value="Tipo")
        ws.cell(row=1, column=2, value="Paso (pulg)")
        ws.cell(row=1, column=3, value="Ancho (pulg)")
        ws.cell(row=1, column=4, value="Resistencia (lbf)")
        ws.cell(row=1, column=5, value="Peso Prom (lbf/pie)")
        ws.cell(row=1, column=6, value="Dia R (pulg)")
        ws.cell(row=1, column=7, value="E.H.M. (pulg)")
        ws.cell(row=1, column=8, value="Longitud de paso")
        ws.cell(row=1, column=9, value="Longitud")
        ws.cell(row=1, column=10, value="Distancia entre centros")

    # Obtener la última fila con datos
    last_row = ws.max_row + 1

    # Datos a guardar en la tabla
    datos = [[Tip, paso, Ancho, Resis, Peso_prom, Diametro_resis, E_H_M, Lp, lon, cen]]

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 22

    # Guardar los datos en la tabla
    for row_data in datos:
        for col_idx, col_data in enumerate(row_data, start=1):
            cell = ws.cell(row=last_row, column=col_idx, value=str(col_data))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell.border + Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        last_row += 1

    # Guardar el archivo de Excel
    wb.save("Tabla.xlsx")

#Funcion para imprimir los encabezadas
def imprimir_encabezado(titulo):
    header = f"""
{'='*80}
===={titulo.center(72)}====
{'='*80}\n"""
    print(header)   


#Inicio del Programa
if __name__ == '__main__':

    while Circuito:
        imprimir_encabezado("Ingreso de datos")

        #Ingreso de los valores de entrada del usuario
        H_nom = float(input("Potencia nominal: "))
        n_d = float(input("Factor de diseño: "))
        while True:
             n = float(input("RPM: "))
             if n <= 3000 :
                 break
             else:
                 print("\nLa velocidad de la catarina inpulsadora tiene que ser mayor de 3000\n")
        tipo = int(input("Tipo de par de torsión:\n1. Par de torsión normal.\n2. Par de torsión alto o no uniforme\nElección: "))
        ks = factor_servicio(tipo)
        while True:
            N1 = float(input("Numero de dientes de la catarina inpulsadora: "))
            if N1 >= 11:
                break
            else:
                print("\nEl numero de dientes de la catarina inpulsadora tiene que ser mayor que 11\n")
        N2 = float(input("Numero de dientes de la catarina inpulsada: "))
        Cp = float(input("Distancia de centro de pasos: "))
        print("\n")

        #Funcion para obtener el k1
        k1 = factor_k1(N1) 
        n1 =[]
        n2 = []

        #Encontrar los Htab para los numeros de hileras
        for i in range(len(TABLA_17_23)):
            k2 = hileras(i)
            Htab = "{:.3f}".format(calculo_Htab_requerido(H_nom, k1, k2, n_d, ks))
            n1.append({"Hileras": i+1 ,"K2": k2, "Htabs":Htab})
        imprimir_tabla(n1)
        print("\n")

        #Encontrar todos los tipos de cadenas que cumplen con las condiciones iniciales
        for i in range(len(n1)):
            valores_tab = H_TAB(TABLA_17_20_A, TABLA_17_20_B, TABLA_17_20_C, TABLA_17_20_C2, n, float(n1[i]["Htabs"]),i+1)
            n2.append(valores_tab)

        #Almacenar los tipos de cademas encontrados
        k = 0
        tabla_grande = []
        for item in n2:
            sub_n2 = item
            for j_item in sub_n2:
                k += 1
                agregado = {"índice": k}
                agregado.update(j_item)
                j_item = agregado
                tabla_grande.append(j_item)
        if tabla_grande == []:
            print(f'No se encontró una cadena adecuada para la potencia {H_nom} y una velocidad de {n}')
        else:
            #imprimir la tabla de almacenamiento
            imprimir_tabla(tabla_grande)
            
            #buscar los valores de la fila seleccionada
            fila = int(input('Fila: '))
            fila-=1
            for i in range(len(tabla_grande)):
                if fila == tabla_grande[i]["índice"]:
                    break
            tipo = tabla_grande[fila]["ANSI"]

            #Extraer los parametros selecionados
            for i in range(len(TABLA_17_19_1)):
                if tipo == TABLA_17_19_1[i]["ANSI"]:
                    parametros = [TABLA_17_19_1[i]]

            #impresion de los valores obtenidos
            imprimir_tabla(parametros)

        Tip = parametros[0]["ANSI"]
        paso = parametros[0]["Paso (pulg)"]
        Ancho = parametros[0]["Ancho (pulg)"]
        Resis = parametros[0]["Resistencia (lbf)"]
        Peso_prom = parametros[0]["Peso Prom (lbf\/pie)"]
        Diametro_resis = parametros[0]["Dia R (pulg)"]
        E_H_M = parametros[0]["E.H.M. (pulg)"]
        Lp = long_paso(Cp, N1, N2)
        lon = longitud(Lp, paso)
        cen = distan_centro(Cp, paso)

        #Guardar los valores obtenidos en un tabla de excel
        guardar_datos_en_tabla(Tip, paso, Ancho, Resis, Peso_prom, Diametro_resis, E_H_M, Lp, lon, cen)
        print('')
        print(f'LONGITUD DE PASOS {Lp} / LONGITUD DE LA CADENA {lon} / DISTANCIA DE CENTROS {cen}')
        

        #Funcion para preguntar la continuidad del programa
        pregunta = True
        while pregunta:
            y =  input('\nDESEA REALIAZR OTRO CALCULO [Y/N]:\n')
            y = y.upper()
            if y == 'N':
                Circuito = False
                pregunta = False
            elif y == 'Y':
                pregunta = False
            else:
                print("\nEL COMANDO INGRESANO NO ES VALIDO")
            
    imprimir_encabezado("MÍNIMO MI 20")
